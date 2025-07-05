import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

class ExcelGenerator:
    def __init__(self, output_dir="excel_generator"):
        """
        Инициализация генератора Excel файлов
        
        Args:
            output_dir (str): Директория для сохранения Excel файлов
        """
        self.output_dir = output_dir
        self._ensure_output_dir()
    
    def _ensure_output_dir(self):
        """Создает директорию для сохранения файлов, если она не существует"""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
    
    def generate_access_codes_excel(self, codes_data):
        """
        Генерирует Excel файл с кодами доступа
        
        Args:
            codes_data (list): Список словарей с данными кодов
                [{'code': '1234-5678-9012', 'hash': '...', 'used': False, 'used_at': None}, ...]
        
        Returns:
            str: Имя созданного файла
        """
        # Создаем новую книгу Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Коды доступа"
        
        # Заголовки
        headers = ['№', 'Код доступа', 'Статус', 'Дата активации', 'Количество использований']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Данные
        for row, code_data in enumerate(codes_data, 2):
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=code_data['code'])
            
            # Статус
            status = 'Активирован' if code_data['used'] else 'Не активирован'
            ws.cell(row=row, column=3, value=status)
            
            # Дата активации
            used_at = code_data.get('used_at')
            if used_at:
                try:
                    # Парсим JSON строку с датой
                    import json
                    date_str = json.loads(used_at)
                    date_obj = datetime.datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                    formatted_date = date_obj.strftime('%d.%m.%Y %H:%M:%S')
                    ws.cell(row=row, column=4, value=formatted_date)
                except:
                    ws.cell(row=row, column=4, value=used_at)
            else:
                ws.cell(row=row, column=4, value='-')
            
            # Количество использований
            usage_count = code_data.get('usage_count', 0)
            ws.cell(row=row, column=5, value=usage_count)
        
        # Автоматическая ширина столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем Excel файл
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'access_codes_{timestamp}.xlsx'
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        
        return filename
    
    def generate_stats_excel(self, stats_data):
        """
        Генерирует Excel файл со статистикой
        
        Args:
            stats_data (dict): Словарь со статистикой
                {'total_codes': 500, 'used_codes': 10, 'unused_codes': 490, 'generated_at': '...'}
        
        Returns:
            str: Имя созданного файла
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Статистика кодов"
        
        # Заголовок
        title_cell = ws.cell(row=1, column=1, value="Статистика кодов доступа")
        title_cell.font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        # Данные статистики
        stats_rows = [
            ['Всего кодов', stats_data.get('total_codes', 0)],
            ['Активировано', stats_data.get('used_codes', 0)],
            ['Не активировано', stats_data.get('unused_codes', 0)],
            ['Процент активации', f"{(stats_data.get('used_codes', 0) / max(stats_data.get('total_codes', 1), 1) * 100):.1f}%"],
            ['Общее количество использований', stats_data.get('total_usage_count', 0)]
        ]
        
        for row, (label, value) in enumerate(stats_rows, 3):
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
        
        # Дата генерации
        generated_at = stats_data.get('generated_at')
        if generated_at:
            try:
                import json
                date_str = json.loads(generated_at)
                date_obj = datetime.datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                formatted_date = date_obj.strftime('%d.%m.%Y %H:%M:%S')
                ws.cell(row=len(stats_rows) + 4, column=1, value="Дата генерации:").font = Font(bold=True)
                ws.cell(row=len(stats_rows) + 4, column=2, value=formatted_date)
            except:
                ws.cell(row=len(stats_rows) + 4, column=1, value="Дата генерации:").font = Font(bold=True)
                ws.cell(row=len(stats_rows) + 4, column=2, value=generated_at)
        
        # Автоматическая ширина столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем файл
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'access_codes_stats_{timestamp}.xlsx'
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        
        return filename
    
    def get_excel_file_path(self, filename):
        """
        Получает полный путь к Excel файлу
        
        Args:
            filename (str): Имя файла
            
        Returns:
            str: Полный путь к файлу
        """
        return os.path.join(self.output_dir, filename)
    
    def list_excel_files(self):
        """
        Возвращает список всех Excel файлов в директории
        
        Returns:
            list: Список имен файлов
        """
        if not os.path.exists(self.output_dir):
            return []
        
        excel_files = []
        for file in os.listdir(self.output_dir):
            if file.endswith('.xlsx'):
                excel_files.append(file)
        
        return sorted(excel_files, reverse=True)  # Сначала новые файлы
    
    def delete_excel_file(self, filename):
        """
        Удаляет Excel файл
        
        Args:
            filename (str): Имя файла для удаления
            
        Returns:
            bool: True если файл удален, False если файл не найден
        """
        filepath = self.get_excel_file_path(filename)
        if os.path.exists(filepath):
            os.remove(filepath)
            return True
        return False 